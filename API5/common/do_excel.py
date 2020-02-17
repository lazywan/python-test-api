#-*-conding:utf-8-*-
#@Time :2020/1/1719:17
#@Author:caowanwan
#@Email:1127825076@qq.com
#@File:do_excel.py
from openpyxl import load_workbook
from openpyxl import workbook
from API5.common.read_config import ReadConfig
from API5.common.read_logging import ReadLog
from API5.common import project_path

class DoExcel:
    '''该类完成数据的读取以及的数据写回'''
    my_log = ReadLog()
    rc=ReadConfig(project_path.config_path)

    def __init__(self,filename,sheetname):
        self.filename=filename  #excel工作簿文件名或地址
        self.sheetname=sheetname


    def read_data(self,option):
        tel =self.get_tel()  #获取的号码是string类型
        case_id=self.rc.get_other(option,'case_id')
        wb=load_workbook(self.filename)
        sheet=wb.get_sheet_by_name(self.sheetname)

        test_data=[]
        self.my_log.info('开始读取数据')
        for i in range(2,sheet.max_row+1):
            row_data = {}
            row_data['CaseId'] = sheet.cell(i, 1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['Title'] = sheet.cell(i, 3).value
            row_data['Url'] = sheet.cell(i, 4).value
            row_data['Method'] = sheet.cell(i, 5).value
            #生成新的手机号防止注册时报错已经注册，利用find和replace函数
            if sheet.cell(i,6).value.find('tel')!=-1: #find函数-1表示没找到，不等于-1表示找到这个字符串
                row_data['Params'] = sheet.cell(i, 6).value.replace('tel',str(tel))
                self.update_phone(int(tel)+1)  #这个tel也要转换为字符串
            else:
                row_data['Params'] = sheet.cell(i, 6).value
            row_data['Sql']= sheet.cell(i, 7).value  #加上去的sql语句
            row_data['ExpectedResult'] = sheet.cell(i,8).value
            self.my_log.info('现在a的值是{}'.format(row_data))
            test_data.append(row_data)
        wb.close()
        finall_data=[] #空列表
        if case_id=='all':#如果case_id=all，执行所有用例，如果是列表，执行列表中指定行数据
            finall_data=test_data  #把测试用例赋值给finall_data变量
        else:
            for i in case_id: #case_id是列表，要遍历
                finall_data.append(test_data[i-1])  # 当i=1表示要执行第一天用例，即test[1-1]=test[0]就是第一条用例
        return finall_data

    def get_tel(self):
        wb = load_workbook(self.filename)
        sheet = wb.get_sheet_by_name('tel')  #这个sheetname要写死
        tel=sheet.cell(1,2).value
        wb.close()
        return tel #返回电话号码的值,且该类型时int,所以上面函数替换的时候，要加str(tel)

    def update_phone(self,new_tel):
        wb = load_workbook(self.filename)
        sheet = wb['tel'] #这个sheetname要写死
        sheet.cell(1,2,new_tel)
        wb.save(self.filename)
        wb.close()

    def write_back(self,row,col,value):
        wb = load_workbook(self.filename)
        sheet = wb[self.sheetname]
        sheet.cell(row=row,column=col).value=value
        wb.save(self.filename)
        wb.close()

if __name__=='__main__':
    ex=DoExcel(project_path.case_path,'add_loan')
    # print(ex.read_excel('test_api.xlsx','register'))
    # ex.create_excel('test_0117.xlsx')
    b=ex.read_data('RegisterCase')
    print(b)
